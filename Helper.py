import secrets
from openpyxl import load_workbook
from datetime import datetime

def generate_unique_glid():
    glid = secrets.token_hex(2)  # Generate a random 4-digit hexadecimal
    while glid_exists(glid):  # Keep generating until a unique GLID is found
        glid = secrets.token_hex(2)
    return glid

def glid_exists(glid):
    wb = load_workbook('glids.xlsx')
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        if row[0] == glid:
            return True
    return False

def save_glid(glid):
    wb = load_workbook('glids.xlsx')
    ws = wb.active
    ws.append([glid, '', '', datetime.now()])
    wb.save('glids.xlsx')

def delete_glid(glid):
    wb = load_workbook('glids.xlsx')
    ws = wb.active
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row[0] == glid:
            ws.delete_rows(idx + 1)  # Adding 1 to the index to get the actual row number
            break
    wb.save('glids.xlsx')
