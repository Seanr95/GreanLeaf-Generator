from flask import Flask, render_template, request, redirect, send_file, jsonify
from openpyxl import load_workbook
from datetime import datetime
import re
import io
import logging
import secrets

app = Flask(__name__, template_folder='templates', static_folder='templates')

# Configure logging
logging.basicConfig(filename='GreebLeafGenerator.log', level=logging.INFO,format='%(asctime)s [%(levelname)s]: %(message)s', datefmt='%Y-%m-%d %H:%M:%S')

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/generate_glid', methods=['GET'])
def generate_glid():
    # Generate a new GLID
    glid = generate_unique_glid()
    # Save the GLID to the Excel file
    save_glid(glid)
     # Add a log message for the generated GLID
    logging.info(f'Generated and saved GLID: {glid}')
    return jsonify({'glid': glid})

@app.route('/add_glid', methods=['POST'])
def add_glid():
    data = request.get_json()
    glid = data['glid']

    # Validate the GLID using the regex pattern
    if not re.match(r'^[0-9a-fA-F]{4}$', glid):
        return jsonify({'error': "Invalid GLID. Please enter a 4-digit hexadecimal value."})

    # Check if GLID already exists
    if not glid_exists(glid):
        # Save the GLID to the Excel file
        save_glid(glid)
         # Add a log message for the Add GLID
        logging.info(f'Added GLID was saved in CSV: {glid}')
    else:
         # Add a log message for the add GLID
        logging.info(f'The following GLID aleady exist in the CSV: {glid}')

    return jsonify({'success': True})

@app.route('/delete_glid', methods=['POST'])
def delete_glid_entry():
    data = request.get_json()
    glid = data['glid']

    # Validate the GLID using the regex pattern
    if not re.match(r'^[0-9a-fA-F]{4}$', glid):
        return jsonify({'error': "Invalid GLID. Please enter a 4-digit hexadecimal value."})

    if not glid_exists(glid):
         # Add a log message for the delete GLID
        logging.info(f'The following GLID doesnt exist for delete: {glid}')
        return jsonify({'error': "GLID does not exist."})

    delete_glid(glid)
     # Add a log message for the delete GLID
    logging.info(f'successfully deleted GLID: {glid}')
    return jsonify({'success': True})

@app.route('/download_glids')
def download_glids():
    return send_file('glids.xlsx', as_attachment=True)

@app.errorhandler(Exception)
def handle_error(e):
    # Log the error message
    logging.error(f'An error occurred: {str(e)}', exc_info=True)
    return jsonify({'error': 'An internal server error occurred.'}), 500

def generate_unique_glid():
    glid = secrets.token_hex(2)  # Generate a random 4-digit hexadecimal
    while glid_exists(glid):  # Keep generating until a unique GLID is found
        glid = secrets.token_hex(2)
    # Add a log message for the generated GLID
    logging.info(f'Generated GLID: {glid}')
    return glid

def glid_exists(glid):
    wb = load_workbook('glids.xlsx')
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        if row[0] == glid:
            return True
    return False

def save_glid(glid):
    wb = load_workbook('glids.xlsx')
    ws = wb.active
    ws.append([glid, '', '', datetime.now()])
    wb.save('glids.xlsx')
    # Add a log message for the saved GLID
    logging.info(f'The following GLID was saved in CSV: {glid}')

def delete_glid(glid):
    wb = load_workbook('glids.xlsx')
    ws = wb.active
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row[0] == glid:
            ws.delete_rows(idx + 1)  # Adding 1 to the index to get the actual row number
            break
    wb.save('glids.xlsx')
    # Add a log message for the Delete GLID
    logging.info(f'Deleted GLID: {glid}')

if __name__ == '__main__':
    # Running the app without debug mode
    app.run(debug=False)
